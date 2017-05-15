import openpyxl
import os
import textwrap
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw 
import argparse


parser = argparse.ArgumentParser()
parser.add_argument('-a', '--all',  action="store_true", help='stores output to an /all directory for easy importing')
args = parser.parse_args()


#Perhaps an enum would have served me better here...
COLUMN_NAME = 1
COLUMN_TYPE = 2
COLUMN_HEALTH = 3
COLUMN_TEXT = 4

CREATURE_NAME_LOCATION = (95, 58)
CREATURE_HEALTH_LOCATION = (40, 280)
CREATURE_TEXT_LOCATION = (70, 530)
CREATURE_RIGHT_BAR_LOCATION = (820,58) #The far right side of the name bar
CREATURE_TOKEN_WIDTH = 64 #Size of a health token

TEXT_RESOLUTION = 42 #Vertical distance between lines
CREATURE_TEXT_WIDTH = 42


TEXT_COLOR = (0,0,0) #black

wb = openpyxl.load_workbook('Elderberry.xlsx')
sheet = wb.get_sheet_by_name('Cards')

creature_name_font = ImageFont.truetype("VerilySerifMono.otf", 48)
creature_text_font = ImageFont.truetype("VerilySerifMono.otf", 32)
costFont = ImageFont.truetype("VerilySerifMono.otf", 32)
textFont = ImageFont.truetype("VerilySerifMono.otf", 16)
textFontLong = ImageFont.truetype("VerilySerifMono.otf", 12)


#Make all folder, if necessary
if args.all:
	if not os.path.exists('all'):
		os.makedirs('all')
		
		
for i in range(2,sheet.max_row + 1):
	folder = sheet.cell(row=i, column=COLUMN_TYPE).value
	
	if not os.path.exists(folder):
		os.makedirs(folder)
	
	#Get the right template
	background = Image.open('templates/' + folder + ".png")
	draw = ImageDraw.Draw(background)

	#Draw type-specific things
	
	#Creature
	if sheet.cell(row=i,column=COLUMN_TYPE).value == "Creature":
	
		#Draw NAME
		draw.text(CREATURE_NAME_LOCATION, str(sheet.cell(row=i,column=COLUMN_NAME).value), TEXT_COLOR, font=creature_name_font)
		
		health = str(sheet.cell(row=i, column=COLUMN_HEALTH).value).split(',')
		
		offset = CREATURE_TOKEN_WIDTH
		
		for x in range(0, len(health)):
			offset = offset - CREATURE_TOKEN_WIDTH
			health[x] = health[x].strip()
			img = Image.open('icons/' + health[x] + '.png')
			
			#Draw each icon 3 times
			for y in range(0,3):
				location = (CREATURE_RIGHT_BAR_LOCATION[0] - CREATURE_TOKEN_WIDTH * x - offset, CREATURE_RIGHT_BAR_LOCATION[1], CREATURE_RIGHT_BAR_LOCATION[0] - CREATURE_TOKEN_WIDTH * (x-1) - offset, CREATURE_RIGHT_BAR_LOCATION[1] + CREATURE_TOKEN_WIDTH)
				offset = offset + CREATURE_TOKEN_WIDTH
				background.paste(img, location)

		img.close()


		#Text wrap
		if sheet.cell(row=i,column=COLUMN_TEXT).value is not None:
			text = textwrap.wrap(str(sheet.cell(row=i,column=COLUMN_TEXT).value), width=CREATURE_TEXT_WIDTH)
			count = 0
			for t in text:
				draw.text((CREATURE_TEXT_LOCATION[0], CREATURE_TEXT_LOCATION[1] + count * TEXT_RESOLUTION), str(t), TEXT_COLOR, font=creature_text_font)
				count = count + 1
		
		
		
		
		
		
		
		
		
		
		
		
		
		
	#Tower
	if sheet.cell(row=i,column=COLUMN_TYPE).value == "Tower":
		draw.text(HEALTH_LOCATION, "Attacks per Round: " + str(sheet.cell(row=i,column=COLUMN_SPEED).value), TEXT_COLOR, font=textFont)
		draw.text(SPEED_LOCATION, "Damage per Attack: " + str(sheet.cell(row=i,column=COLUMN_DAMAGE).value), TEXT_COLOR, font=textFont)
		
		#Text wrap
		if sheet.cell(row=i,column=COLUMN_TEXT).value is not None:
			text = textwrap.wrap(str(sheet.cell(row=i,column=COLUMN_TEXT).value), width=SHEET_WIDTH)
			myfont = textFont
			if len(text) > 4:#Max rows that fit on a sheet (after tower info)
				myfont = textFontLong
				text = textwrap.wrap(sheet.cell(row=i,column=COLUMN_TEXT).value, width=SHEET_WIDTH + 10)
			
			count = 0
			for t in text:
				draw.text((CREATURE_TEXT_LOCATION[0], CREATURE_TEXT_LOCATION[1] + count * TEXT_RESOLUTION), str(t), TEXT_COLOR, font=myfont)
				count = count + 1
			
	#Spell
	if sheet.cell(row=i,column=COLUMN_TYPE).value == "Spell":
		if sheet.cell(row=i,column=COLUMN_TEXT).value is not None:
			text = textwrap.wrap(sheet.cell(row=i,column=COLUMN_TEXT).value, width=SHEET_WIDTH)
			myfont = textFont
			if len(text) > 8:#Max rows that fit on a sheet
				myfont = textFontLong#Could make this a sanitize_text function
				text = textwrap.wrap(sheet.cell(row=i,column=COLUMN_TEXT).value, width=SHEET_WIDTH + 10)
			
			count = 0
			for t in text:
				draw.text((HEALTH_LOCATION[0], HEALTH_LOCATION[1] + count * TEXT_RESOLUTION), str(t), TEXT_COLOR, font=myfont)
				count = count + 1
				
			
	background.save(folder + '/' + sheet.cell(row=i,column=COLUMN_NAME).value.replace(" ", "_") + '.png')
	
	#make a second copy for easy import
	if args.all:
		background.save('all' + '/' + sheet.cell(row=i,column=COLUMN_NAME).value.replace(" ", "_") + '.png')
		
	background.close()